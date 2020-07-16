# coding = utf-8
'''
封装操作Excel的类OperateExcle


    # 获得行数
    def get_row(self):
        wb =openpyxl.load_workbook(self.filename)
        ws= wb.active
        return ws.max_row

    # 获得列数
    def get_col(self):

        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        return ws.max_column

'''

import openpyxl

class OperateExcle:

    # 初始化获得sheet 表
    def __init__(self,filename=None):
        if filename:
            self.filename= filename
        else:
            self.filename=r"C:\Users\fan\Desktop\测试用例模板.xlsx"
        try:
            wb = openpyxl.load_workbook(self.filename)
            self.ws = wb.active
        except:
            print('获取Excel失败')
        # self.get_row = self.get_row           调试用
        # self.get_col = self.get_col           调试用
        # self.ws = self.open_ws()              调试用

    # def open_ws(self):                        调试用
    #     wb = openpyxl.load_workbook(self.filename)
    #     return wb.active

    def get_row(self):                          # 根据初始化表self.ws 获得行数
        return self.ws.max_row

    def get_col(self):
        return self.ws.max_column

    def get_value(self,row,col):
        return self.ws.cell(row=row,column=col).value
if __name__ == '__main__':
    filename = r"C:\Users\fan\Desktop\用例0.xlsx"
    case = OperateExcle(filename)
    print(case.get_col(),case.get_row())



