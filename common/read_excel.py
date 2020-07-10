# coding = utf-8

'''
读取 excel 用例方法
1.打开工作薄：openxl.load_workbook
2.读取工作表 .active(默认读取第一张表)
3.获得工作表中最大的行数.max_row
4.遍历最大行数，根据行数获得每行单元格的数据，存储起来 .cell(row=,column=).value
5.获得数据根据每条用例为单位，转化成字典 dict（）
6.把字典放到列表中 .append()
'''
import openpyxl

def open_excel(path):

    wb = openpyxl.load_workbook(path)
    return wb.active
'''
    方法1
    直接读取exccel 用例 ，返回列表嵌套字典数据类型
'''
def read_excel(path):
    # path = r"C:\Users\fan\Desktop\个人资料表.xlsx"
    ws = open_excel(path)
    case = ws.max_row                         # 获得行数
    case_list=[]
    for row in range(2, case+1):                     # 获得每个用例的数据
        dict_case = dict(                               # 转成字典
        id = ws.cell(row=row, column=1).value,
        modle = ws.cell(row=row, column=2).value,
        case_name = ws.cell(row=row, column=3).value,
        interface_name = ws.cell(row=row, column=4).value,
        url = ws.cell(row=row, column=5).value,
        method = ws.cell(row=row, column=6).value,
        header = ws.cell(row=row, column=7).value,
        parmas = ws.cell(row=row,column=8).value,
        executeOrNot = ws.cell(row=row,column=9).value,
        expectRes = ws.cell(row=row,column=10).value)
        case_list.append(dict_case)
    return case_list
'''
    方法2
    get_cxcutre_case() 返回执行列为yes的用例数据。
    需要传2个参数，用例的路径，是否执行的列名称，如“A”是模板中的执行列，excColName=“A”
'''
def get_excutre_case(path, excColName):
    ws = open_excel(path)
    execute_col = ws[excColName]
    exe = []
    list_case = []
    # print(execute_col)
    for i in execute_col[1:]:
        # print(i.coordinate,i.value)
        if i.value == 'yes':
            exe.append(i.row)

    for case in exe:
        dictc_case = dict(
            id=ws.cell(row=case, column=1).value,
            mode=ws.cell(row=case, column=2).value,
            case_name=ws.cell(row=case, column=3).value,
            api_name=ws.cell(row=case, column=4).value,
            url=ws.cell(row=case, column=5).value,
            method=ws.cell(row=case, column=6).value,
            head=ws.cell(row=case, column=7).value,
            parmer=ws.cell(row=case, column=8).value,
            cause_or_not=ws.cell(row=case, column=9).value,
            expected_result=ws.cell(row=case, column=10).value)
        list_case.append(dictc_case)
    return list_case
if __name__ == '__main__':
    path = r"C:\Users\fan\Desktop\个人资料表.xlsx"
    # print(read_excel(path))
    # print(get_excutre_case(path, "I"))